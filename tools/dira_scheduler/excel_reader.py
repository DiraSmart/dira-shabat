"""Parse Dira Shabat scheduler Excel files."""
from __future__ import annotations

import warnings as _warnings
from dataclasses import dataclass
from datetime import time as _dt_time
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


SUPPORTED_DOMAINS = frozenset({
    "light", "climate", "switch", "fan",
    "cover", "media_player", "input_boolean",
})


class DispositivosError(ValueError):
    """Raised when the Dispositivos sheet is missing or malformed."""


@dataclass(frozen=True)
class Device:
    area: str
    nombre: str
    domain: str
    entity_id: str


def read_dispositivos(path: str | Path) -> dict[tuple[str, str], Device]:
    """Load the Dispositivos sheet into an (Area, Nombre) -> Device map."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Dispositivos" not in wb.sheetnames:
        raise DispositivosError("Sheet 'Dispositivos' is missing from the workbook")
    ws = wb["Dispositivos"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise DispositivosError("Sheet 'Dispositivos' is empty")
    # Skip header row
    out: dict[tuple[str, str], Device] = {}
    for idx, row in enumerate(rows[1:], start=2):
        area, nombre, tipo, entity_id, *_ = (list(row) + [None] * 5)[:5]
        if not any((area, nombre, tipo, entity_id)):
            continue  # blank row
        if not (area and nombre and tipo and entity_id):
            raise DispositivosError(
                f"Row {idx} in Dispositivos has missing required field"
            )
        area, nombre, tipo, entity_id = (
            str(area).strip(), str(nombre).strip(),
            str(tipo).strip(), str(entity_id).strip(),
        )
        if tipo not in SUPPORTED_DOMAINS:
            raise DispositivosError(
                f"Row {idx}: unknown domain '{tipo}'. "
                f"Supported: {', '.join(sorted(SUPPORTED_DOMAINS))}"
            )
        key = (area, nombre)
        if key in out:
            raise DispositivosError(
                f"Row {idx}: duplicate (Area, Nombre) entry: {key}"
            )
        out[key] = Device(area=area, nombre=nombre, domain=tipo, entity_id=entity_id)
    return out


@dataclass(frozen=True)
class ScheduleCell:
    time: str               # "HH:MM"
    in_erev_band: bool
    area: str
    nombre: str
    value: str | int


_EREV_SENTINEL = "erev shabat"


def _format_time(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, _dt_time):
        return value.strftime("%H:%M")
    s = str(value).strip()
    if not s:
        return None
    # Accept "HH:MM" (allow single-digit hour)
    parts = s.split(":")
    if len(parts) != 2:
        return None
    try:
        h, m = int(parts[0]), int(parts[1])
    except ValueError:
        return None
    if not (0 <= h < 24 and 0 <= m < 60):
        return None
    return f"{h:02d}:{m:02d}"


def read_schedule_sheet(path: str | Path, sheet_name: str) -> Iterator[ScheduleCell]:
    """Yield ScheduleCell for every non-empty cell in a schedule sheet.

    Returns nothing if the sheet doesn't exist.
    Row 1 = area headers (col A blank), row 2 = nombre headers.
    Column A is times; a row with col A == 'Erev Shabat' (case-insensitive)
    starts the Erev band; the band ends on the next row whose col A is empty
    or non-time.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return

    # Build column -> (area, nombre) mapping
    area_row = list(rows[0])
    nombre_row = list(rows[1])
    col_map: dict[int, tuple[str, str]] = {}
    last_area: str | None = None
    for col_idx, area in enumerate(area_row):
        if col_idx == 0:
            continue
        if area is not None:
            last_area = str(area).strip() or last_area
        nombre = nombre_row[col_idx] if col_idx < len(nombre_row) else None
        if last_area and nombre:
            col_map[col_idx] = (last_area, str(nombre).strip())

    in_erev = False
    for row in rows[2:]:
        col_a = row[0]
        if col_a is not None and str(col_a).strip().lower() == _EREV_SENTINEL:
            in_erev = True
            continue
        time = _format_time(col_a)
        if time is None:
            # Invalid time (non-empty col_a that doesn't parse) → warn.
            if col_a is not None and str(col_a).strip():
                _warnings.warn(
                    f"Sheet {sheet_name!r}: invalid time in column A: "
                    f"{col_a!r}; row skipped.",
                    UserWarning,
                    stacklevel=2,
                )
            in_erev = False  # band ends at first non-time row after entering
            continue
        for col_idx, (area, nombre) in col_map.items():
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            if isinstance(value, str):
                s = value.strip()
                # Coerce numeric strings to int
                try:
                    value = int(s)
                except ValueError:
                    value = s
            elif isinstance(value, (int, float)):
                value = int(value)
            yield ScheduleCell(
                time=time,
                in_erev_band=in_erev,
                area=area,
                nombre=nombre,
                value=value,
            )
