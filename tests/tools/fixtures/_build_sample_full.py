"""Run once to (re)create sample_full.xlsx."""
from pathlib import Path

from openpyxl import Workbook


def build():
    wb = Workbook()
    wb.remove(wb.active)

    disp = wb.create_sheet("Dispositivos")
    disp.append(["Area", "Nombre", "Tipo", "entity_id", "Acepta número"])
    disp.append(["Sala", "Spots", "light", "light.sala_spots", "dimming %"])
    disp.append(["Sala", "A/C", "climate", "climate.sala_ac", "temperatura"])
    disp.append(["Cocina", "Luz", "light", "light.cocina_main", "dimming %"])

    encasa = wb.create_sheet("En Casa")
    encasa.append([None, "Sala", "Sala", "Cocina"])
    encasa.append([None, "Spots", "A/C", "Luz"])
    encasa.append(["Erev Shabat", None, None, None])
    encasa.append(["18:00", "ON", None, None])
    encasa.append(["18:30", None, 22, None])
    encasa.append([None, None, None, None])  # ends Erev band
    encasa.append(["19:00", None, None, "ON"])
    encasa.append(["20:00", 30, None, None])
    encasa.append(["23:00", "OFF", "OFF", "OFF"])

    fuera = wb.create_sheet("Fuera")
    fuera.append([None, "Cocina"])
    fuera.append([None, "Luz"])
    fuera.append(["19:00", "ON"])
    fuera.append(["23:00", "OFF"])

    diario = wb.create_sheet("Diario Lun-Vier")
    diario.append([None, "Sala"])
    diario.append([None, "Spots"])
    diario.append(["07:00", "ON"])
    diario.append(["22:30", "OFF"])

    out = Path(__file__).parent / "sample_full.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    build()
